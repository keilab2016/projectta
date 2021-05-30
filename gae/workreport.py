import gspread
from oauth2client.service_account import ServiceAccountCredentials
import openpyxl
import re
import sys
import os
import tempfile

def getAllData(spreadsheet_key, sheet_name):
    #jsonファイルを使って認証情報を取得
    scope = ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
    c = ServiceAccountCredentials.from_json_keyfile_name('projectta2021-d0f50421fa21.json', scope)

    #認証情報を使ってスプレッドシートの操作権を取得
    gs = gspread.authorize(c)

    #共有したスプレッドシートのキー（後述）を使ってシートの情報を取得
    worksheet = gs.open_by_key(spreadsheet_key).worksheet(sheet_name)
    list_of_lists = worksheet.get_all_values()

    #print(list_of_lists)
    return list_of_lists

def getAllID(list_of_lists):
    for line in list_of_lists:
        if line[0] == '学籍番号':
            ids = line
    ret=[]
    for e in ids[1:]:
        if e.isnumeric():
            ret.append(int(e))
        else:
            break
    return ret

def findData(userid, month, list_of_lists):
    alldata={}
    for line in list_of_lists:
        if line[0] == '学籍番号':
            ids = line
        elif '氏名' in line[0] or '名前' in line[0]:
            names = line
        elif '研究室' in line[0]:
            labs = line
        elif '時給' in line[0] or '単価' in line[0]:
            price = line
        else:
            result = re.findall(r'([0-9]+)/([0-9]+)', line[0])
            if len(result)==1 and int(result[0][0])==month:
                alldata[result[0][0]+'/'+result[0][1]]=line

    #print(ids)
    #print(names)
    #print(alldata)
    data={}
    if userid in ids:
        idx = ids.index(userid)
        data['id']=ids[idx]
        data['name']=names[idx]
        data['lab']=labs[idx]
        data['price']=price[idx]
        for day,line in alldata.items():
            data[day]=line[idx]
    else:
        print(userid,'not found in',ids)
    return data

def writeData(userid, month, data):
    wb = openpyxl.load_workbook('workTableA.xlsx')
    ws = wb.worksheets[0]
    # A7 年月
    ws["A7"].value="令和 3年 " + str(month) + "月"
    # B9 1日の業務内容
    # D9 1日の勤務時間
    # E9 1日の実働時間
    # F9 1日の休憩時間
    # B10 2日の業務内容
    # B24 16日の業務内容
    # I9 17日の業務内容
    # L9
    # M9
    # N9
    # I10 18日の業務内容
    # I23 31日の業務内容
    sum=0.0
    for d,h in data.items():
        if re.match(r'[0-9]+/[0-9]+',d):
            day=int(re.sub(r'[0-9]+/','',d))
            if h == '':
                hour = 0.0
            else:
                hour=float(h)
            endhour=15+int(hour)
            if int(hour)==hour:
                endmin=0
            else:
                endmin=30
            hourstr="15:00\n{0:02d}:{1:02d}".format(endhour,endmin)
            if hour > 0.0:
                if day < 17:
                    ws.cell(row=8+day, column=2, value='プロジェクト学習技術指導補助')
                    ws.cell(row=8+day, column=4, value=hourstr)
                    ws.cell(row=8+day, column=5, value=hour)
                    sum += hour
                else:
                    ws.cell(row=day-8, column=9, value='プロジェクト学習技術指導補助')
                    ws.cell(row=day-8, column=12, value=hourstr)
                    ws.cell(row=day-8, column=13, value=hour)
                    sum += hour
    # C27 氏名
    ws["C27"].value=data['name']
    # C28 学籍番号
    ws["C28"].value=int(data['id'])
    # L25 合計時間
    ws["L25"].value=sum
    # K27 時給
    ws["K27"].value=int(data['price'])
    # K28 合計金額
    ws["K28"].value=sum*int(data['price'])
    outpath=tempfile.gettempdir() + '/' + userid + '.xlsx'
    if os.path.exists(outpath):
        os.remove(outpath)
    wb.save(outpath)
    return outpath

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print('Usage: ' + sys.argv[0] + ' student-id month')
    elif not sys.argv[1].isnumeric() or not sys.argv[2].isnumeric():
        print('Usage: ' + sys.argv[0] + ' student-id month')
    else:
        id = sys.argv[1]
        month = int(sys.argv[2])
        list_of_lists = getAllData('102Fbm3HKFOeCg6Q0CEAQE88CLQ3xvtW3bTPDGydx02A',"2021前期")
        data=findData(id, month, list_of_lists)
        writeData(id, month, data)
